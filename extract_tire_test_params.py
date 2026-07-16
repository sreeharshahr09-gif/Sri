"""
Extract tire test parameters from Excel test files across an entire folder
tree, writing one consolidated file per subfolder.

Layout expected in each source file (on the first sheet):
    A2:A21    -> parameter names
    B2:B21    -> parameter values
The file name (without extension) is used as the column label.

Which files are considered:
    Only Excel files inside subfolders named "input" (case-insensitive) are
    targeted; anything elsewhere in the tree is ignored.

Passing criteria:
    A file is valid only if its parameter names in A2:A10 match the expected
    set. The file name itself is not checked. The expected set is either
    EXPECTED_PARAMS (if filled in) or, otherwise, the A2:A10 signature shared
    by the majority of the targeted files.

How to use:
    1. Keep this script anywhere you like (only ONE copy needed).
    2. Run it:   python extract_tire_test_params.py
    3. A folder-picker window opens -> choose the PARENT folder that holds
       the subfolders of Excel test files.
    4. For every folder in the tree that contains valid test files, a
       "tire_test_parameters.xlsx" is created IN that same folder.

Each output file has two sheets:
    * "Tire Test Parameters" - one column per test case, one row per parameter.
    * "Report"               - per-file Success/Failure status and totals.
"""

import os
import sys
from collections import Counter
from openpyxl import load_workbook, Workbook

# Rows to read (1-based, matching Excel). A2:A21 / B2:B21 -> rows 2..21.
FIRST_DATA_ROW = 2
LAST_DATA_ROW = 21

# Only folders with this name (case-insensitive) are searched for test
# files; Excel files anywhere else in the tree are ignored.
TARGET_FOLDER_NAME = "input"

# Passing criteria: the parameter names in this range must match the
# expected set for the file to count as valid.
KEY_FIRST_ROW = 2
KEY_LAST_ROW = 10

# The expected parameter names for A2:A10, in order. Comparison is
# case-insensitive and ignores surrounding spaces.
#   * Leave this list EMPTY to auto-derive the reference by majority vote
#     across all files under the parent folder.
#   * Or fill in the 9 names explicitly to enforce a fixed master list.
EXPECTED_PARAMS = []

OUTPUT_FILENAME = "tire_test_parameters.xlsx"


def choose_folder():
    """Open a folder-picker dialog and return the selected path (or None)."""
    try:
        import tkinter as tk
        from tkinter import filedialog
    except Exception:
        # tkinter unavailable (rare) -> fall back to the current directory.
        print("Folder picker unavailable; using the current folder.")
        return os.getcwd()

    root = tk.Tk()
    root.withdraw()          # hide the empty root window
    root.attributes("-topmost", True)  # bring dialog to the front
    folder = filedialog.askdirectory(
        title="Select the PARENT folder containing the test subfolders")
    root.destroy()
    return folder or None    # returns "" if the user cancels


def show_message(title, text):
    """Show a pop-up so the result is visible even when double-clicked.

    Falls back to console output + an Enter-to-close pause if tkinter or a
    display is unavailable.
    """
    print(title)
    print(text)
    try:
        import tkinter as tk
        from tkinter import messagebox
        root = tk.Tk()
        root.withdraw()
        root.attributes("-topmost", True)
        messagebox.showinfo(title, text)
        root.destroy()
    except Exception:
        try:
            input("\nPress Enter to close...")
        except Exception:
            pass


def read_file(path):
    """Read sheet 1 of one Excel file.

    Returns (label, key_names, params) where:
        label      -> file name without extension (used as the column header)
        key_names  -> the 9 A2:A10 parameter names (blanks kept as "")
        params     -> [(name, value), ...] for every filled row in A2:A21
    """
    wb = load_workbook(path, data_only=True)
    ws = wb.worksheets[0]  # sheet 1

    label = os.path.splitext(os.path.basename(path))[0]

    key_names = []
    for row in range(KEY_FIRST_ROW, KEY_LAST_ROW + 1):
        v = ws.cell(row=row, column=1).value
        key_names.append("" if v is None else str(v).strip())

    params = []
    for row in range(FIRST_DATA_ROW, LAST_DATA_ROW + 1):
        name = ws.cell(row=row, column=1).value   # column A
        value = ws.cell(row=row, column=2).value  # column B
        if name is None or str(name).strip() == "":
            continue
        params.append((str(name).strip(), value))

    wb.close()
    return str(label).strip(), key_names, params


def _norm(names):
    """Normalise a list of names for case/space-insensitive comparison."""
    return [n.strip().lower() for n in names]


def validate_key_names(key_names, reference):
    """Check A2:A10 names against the expected `reference` list.

    Returns (is_valid, details_message).
    """
    # Every A2:A10 cell must be filled first.
    blanks = [KEY_FIRST_ROW + i for i, n in enumerate(key_names) if n == ""]
    if blanks:
        return False, ("Missing parameter name(s) in row(s): "
                       + ", ".join(str(r) for r in blanks))

    if reference is None:
        # No reference could be established (no complete files at all).
        return True, "%d parameter(s)" % len(key_names)

    if _norm(key_names) != _norm(reference):
        diffs = ["row %d: '%s' != '%s'" % (KEY_FIRST_ROW + i, key_names[i],
                                           reference[i] if i < len(reference) else "")
                 for i in range(len(key_names))
                 if i >= len(reference) or _norm([key_names[i]]) != _norm([reference[i]])]
        return False, "Parameter names differ from expected (" + \
                      "; ".join(diffs) + ")"

    return True, "%d parameter(s)" % len(key_names)


def list_candidate_files(folder):
    """Return the Excel files in one folder, ignoring temp files and output."""
    files = []
    for name in sorted(os.listdir(folder)):
        if not name.lower().endswith((".xlsx", ".xlsm", ".xls")):
            continue
        if name.startswith("~$"):          # Excel lock/temp files
            continue
        if name == OUTPUT_FILENAME:        # don't re-ingest our own output
            continue
        files.append(os.path.join(folder, name))
    return files


def determine_reference(folder_entries):
    """Pick the expected A2:A10 name set.

    Uses EXPECTED_PARAMS if provided; otherwise the A2:A10 signature shared by
    the most files (majority vote). Returns a list of names, or None.
    """
    if EXPECTED_PARAMS:
        return list(EXPECTED_PARAMS)

    counts = Counter()
    example = {}  # normalised signature -> original-case names (first seen)
    for entries in folder_entries.values():
        for _path, result, _exc in entries:
            if result is None:
                continue
            _label, key_names, _params = result
            if any(n == "" for n in key_names):
                continue  # only complete A2:A10 rows can define the reference
            sig = tuple(_norm(key_names))
            counts[sig] += 1
            example.setdefault(sig, list(key_names))

    if not counts:
        return None
    best_sig = counts.most_common(1)[0][0]
    return example[best_sig]


def build_and_write(folder, entries, reference):
    """Validate cached entries for one folder and write its output workbook.

    Returns (n_success, n_failure).
    """
    cases = []            # list of (label, {param_name: value})
    param_order = []      # preserves the order parameters first appear
    report_rows = []      # list of (file_name, status, details)

    for path, result, exc in entries:
        fname = os.path.basename(path)
        if result is None:
            report_rows.append((fname, "Failure", str(exc)))
            continue

        label, key_names, params = result
        is_valid, details = validate_key_names(key_names, reference)
        if not is_valid:
            report_rows.append((fname, "Failure", details))
            continue

        param_map = {}
        for name, value in params:
            if name not in param_order:
                param_order.append(name)
            param_map[name] = value

        cases.append((label, param_map))
        report_rows.append((fname, "Success", details))

    n_success = len(cases)
    n_failure = len(report_rows) - n_success

    _write_output(folder, cases, param_order, report_rows, n_success, n_failure)
    return n_success, n_failure


def _write_output(folder, cases, param_order, report_rows, n_success, n_failure):
    """Write the consolidated data sheet and the report sheet to disk."""
    out_wb = Workbook()

    # --- Sheet 1: consolidated parameters ---
    data_ws = out_wb.active
    data_ws.title = "Tire Test Parameters"

    data_ws.cell(row=1, column=1, value="Parameter")
    for col, (label, _) in enumerate(cases, start=2):
        data_ws.cell(row=1, column=col, value=label)

    for r, param_name in enumerate(param_order, start=2):
        data_ws.cell(row=r, column=1, value=param_name)
        for col, (_, param_map) in enumerate(cases, start=2):
            data_ws.cell(row=r, column=col, value=param_map.get(param_name))

    # --- Sheet 2: report ---
    rep_ws = out_wb.create_sheet("Report")
    rep_ws.cell(row=1, column=1, value="Extraction report for folder:")
    rep_ws.cell(row=1, column=2, value=folder)

    rep_ws.cell(row=3, column=1, value="File Name")
    rep_ws.cell(row=3, column=2, value="Status")
    rep_ws.cell(row=3, column=3, value="Details")
    for i, (fname, status, details) in enumerate(report_rows, start=4):
        rep_ws.cell(row=i, column=1, value=fname)
        rep_ws.cell(row=i, column=2, value=status)
        rep_ws.cell(row=i, column=3, value=details)

    summary_row = 4 + len(report_rows) + 1
    rep_ws.cell(row=summary_row, column=1, value="Total files")
    rep_ws.cell(row=summary_row, column=2, value=len(report_rows))
    rep_ws.cell(row=summary_row + 1, column=1, value="Success")
    rep_ws.cell(row=summary_row + 1, column=2, value=n_success)
    rep_ws.cell(row=summary_row + 2, column=1, value="Failure")
    rep_ws.cell(row=summary_row + 2, column=2, value=n_failure)

    out_wb.save(os.path.join(folder, OUTPUT_FILENAME))


def run():
    """Do the extraction and return a human-readable summary string."""
    parent = choose_folder()
    if not parent:
        return "No folder selected. Nothing to do."

    # --- Pass 1: read every candidate file once, cache the result. ---
    folder_entries = {}  # folder -> [(path, (label, key_names, params) | None, exc | None)]
    for dirpath, _dirnames, _filenames in os.walk(parent):
        # Only look inside folders named "input".
        if os.path.basename(os.path.normpath(dirpath)).lower() != TARGET_FOLDER_NAME:
            continue
        candidates = list_candidate_files(dirpath)
        if not candidates:
            continue
        entries = []
        for path in candidates:
            try:
                entries.append((path, read_file(path), None))
            except Exception as exc:  # unreadable file
                entries.append((path, None, exc))
        folder_entries[dirpath] = entries

    if not folder_entries:
        return ("No '%s' folders with Excel files were found under:\n%s\n\n"
                "Check that your test files sit inside subfolders named '%s'."
                % (TARGET_FOLDER_NAME, parent, TARGET_FOLDER_NAME))

    # --- Determine the expected A2:A10 parameter names. ---
    reference = determine_reference(folder_entries)
    if reference:
        print("Expected A2:A10 parameters:", ", ".join(reference))

    # --- Pass 2: validate and write one output workbook per folder. ---
    folders_done = 0
    grand_success = 0
    grand_failure = 0
    for dirpath, entries in folder_entries.items():
        n_success, n_failure = build_and_write(dirpath, entries, reference)
        folders_done += 1
        grand_success += n_success
        grand_failure += n_failure
        print("Processed:", dirpath,
              "-> success: %d, failure: %d" % (n_success, n_failure))

    return ("Done. {0} '{1}' folder(s) processed.\n"
            "Total files -> success: {2}, failure: {3}\n\n"
            "An output file '{4}' was written into each '{1}' folder."
            .format(folders_done, TARGET_FOLDER_NAME,
                    grand_success, grand_failure, OUTPUT_FILENAME))


def main():
    try:
        summary = run()
        show_message("Tire Test Extractor", summary)
    except Exception:
        import traceback
        show_message("Tire Test Extractor - ERROR", traceback.format_exc())


if __name__ == "__main__":
    main()
